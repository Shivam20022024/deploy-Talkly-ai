import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _auto_adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

def _apply_header_style(ws, header_row=1):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

def _safe_row(row):
    return [str(item) if isinstance(item, (dict, list)) else item for item in row]

def generate_multi_sheet_report(calls_data):
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Colors
    color_hot = "FF9999" # Light Red
    color_warm = "FFCC99" # Light Orange
    color_cold = "99CCFF" # Light Blue
    
    # ---------------------------------------------------------
    # SHEET 1: EXECUTIVE DASHBOARD
    # ---------------------------------------------------------
    ws1 = wb.create_sheet("EXECUTIVE DASHBOARD")
    ws1.append(["Key Performance Indicator", "Value"])
    
    total_leads = len(calls_data)
    hot_leads = sum(1 for c in calls_data if c.get("analysis", {}).get("lead_temperature") == "Hot")
    warm_leads = sum(1 for c in calls_data if c.get("analysis", {}).get("lead_temperature") == "Warm")
    cold_leads = sum(1 for c in calls_data if c.get("analysis", {}).get("lead_temperature") == "Cold")
    
    total_intent = sum(c.get("analysis", {}).get("intent_score", 0) for c in calls_data)
    avg_intent = round(total_intent / total_leads, 2) if total_leads else 0
    
    total_conv = sum(c.get("analysis", {}).get("conversion_probability", 0) for c in calls_data)
    avg_conv = round(total_conv / total_leads, 2) if total_leads else 0
    
    # Missing fields for dashboard we can estimate or leave placeholder
    calls_this_week = total_leads # Simplified
    missed_follow_ups = 0
    
    data1 = [
        ["Total Leads", total_leads],
        ["Hot Leads", hot_leads],
        ["Warm Leads", warm_leads],
        ["Cold Leads", cold_leads],
        ["Avg Buyer Intent %", avg_intent],
        ["Conversion Probability %", avg_conv],
        ["Calls This Week", calls_this_week],
        ["Missed Follow-ups", missed_follow_ups]
    ]
    
    for row in data1:
        ws1.append(_safe_row(row))
        
    _apply_header_style(ws1)
    _auto_adjust_columns(ws1)
    
    # ---------------------------------------------------------
    # SHEET 2: LEAD PIPELINE
    # ---------------------------------------------------------
    ws2 = wb.create_sheet("LEAD PIPELINE")
    headers2 = [
        "Lead ID", "Customer Name", "Phone Number", "Project Interested", 
        "Budget", "Property Type", "Preferred Location", "Lead Temperature",
        "Buyer Intent %", "Conversion Probability %", "Last Call Date", 
        "Next Follow-up Date", "Assigned Agent", "Call Outcome", "Priority"
    ]
    ws2.append(headers2)
    
    for c in calls_data:
        analysis = c.get("analysis") or {}
        reqs = analysis.get("property_requirements") or {}
        temp = analysis.get("lead_temperature", "Warm")
        intent = analysis.get("intent_score", 0)
        
        priority = "High" if temp == "Hot" else "Medium" if temp == "Warm" else "Low"
        dt = c.get("created_at")
        dt_str = dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(dt, datetime) else str(dt)
        
        row = [
            c.get("call_id", ""),
            c.get("customer_name", "Phone Lead"),
            c.get("customer_id", ""),
            reqs.get("location", "N/A"), # Assuming location is project interested
            reqs.get("budget", "N/A"),
            reqs.get("propertyType", "N/A"),
            reqs.get("location", "N/A"),
            temp,
            intent,
            analysis.get("conversion_probability", 0),
            dt_str,
            "TBD", # Next Follow-up Date
            c.get("agent_name", "AI Agent"),
            analysis.get("summary", ""),
            priority
        ]
        ws2.append(_safe_row(row))
        
        # Color coding
        last_row = ws2.max_row
        if temp == "Hot":
            fill = PatternFill(start_color=color_hot, end_color=color_hot, fill_type="solid")
        elif temp == "Warm":
            fill = PatternFill(start_color=color_warm, end_color=color_warm, fill_type="solid")
        else:
            fill = PatternFill(start_color=color_cold, end_color=color_cold, fill_type="solid")
            
        for cell in ws2[last_row]:
            cell.fill = fill

    _apply_header_style(ws2)
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = "A2"
    _auto_adjust_columns(ws2)

    # ---------------------------------------------------------
    # SHEET 3: AGENT PERFORMANCE
    # ---------------------------------------------------------
    ws3 = wb.create_sheet("AGENT PERFORMANCE")
    headers3 = [
        "Agent Name", "Calls Handled", "Avg Talk Time", "Hot Leads Generated", 
        "Conversions Closed", "Follow-up Completion %", "Avg Sentiment Score"
    ]
    ws3.append(headers3)
    
    agents = {}
    for c in calls_data:
        agent = c.get("agent_name", "AI Agent")
        if agent not in agents:
            agents[agent] = {"calls": 0, "hot": 0, "conv": 0, "sentiments": []}
        
        agents[agent]["calls"] += 1
        analysis = c.get("analysis") or {}
        if analysis.get("lead_temperature") == "Hot":
            agents[agent]["hot"] += 1
        
        agents[agent]["conv"] += 1 if analysis.get("conversion_probability", 0) > 80 else 0
        sentiment = c.get("sentiment", "neutral")
        score = 1 if sentiment == "positive" else (-1 if sentiment == "negative" else 0)
        agents[agent]["sentiments"].append(score)
        
    for agent, stats in agents.items():
        avg_sent = round(sum(stats["sentiments"])/len(stats["sentiments"]), 2) if stats["sentiments"] else 0
        ws3.append(_safe_row([
            agent, stats["calls"], "N/A", stats["hot"], 
            stats["conv"], "100%", avg_sent
        ]))
        
    _apply_header_style(ws3)
    ws3.auto_filter.ref = ws3.dimensions
    _auto_adjust_columns(ws3)

    # ---------------------------------------------------------
    # SHEET 4: CALL INTELLIGENCE SUMMARY
    # ---------------------------------------------------------
    ws4 = wb.create_sheet("CALL INTELLIGENCE SUMMARY")
    headers4 = [
        "Customer Name", "Sentiment", "Budget Mentioned", "Property Type", 
        "Preferred Location", "Timeline to Purchase", "Loan Requirement", 
        "Site Visit Interest", "Key Objections", "Competitor Mentioned", 
        "Interest Score", "Language"
    ]
    ws4.append(headers4)
    
    for c in calls_data:
        analysis = c.get("analysis") or {}
        reqs = analysis.get("property_requirements") or {}
        objections = analysis.get("objections") or []
        obj_str = ", ".join([o.get("content", "") for o in objections if isinstance(o, dict)]) if objections else "None"
        
        intents = analysis.get("intents") or []
        site_visit = "Yes" if "Site Visit" in str(intents) else "No"
        
        ws4.append(_safe_row([
            c.get("customer_name", "Phone Lead"),
            c.get("sentiment", "neutral"),
            reqs.get("budget", "N/A"),
            reqs.get("propertyType", "N/A"),
            reqs.get("location", "N/A"),
            reqs.get("timeline", "N/A"),
            "Yes" if reqs.get("loanRequired") else "No",
            site_visit,
            obj_str,
            "N/A", # Competitor Not directly available
            analysis.get("intent_score", 0),
            c.get("language", "English")
        ]))
    _apply_header_style(ws4)
    ws4.auto_filter.ref = ws4.dimensions
    ws4.freeze_panes = "A2"
    _auto_adjust_columns(ws4)

    # ---------------------------------------------------------
    # SHEET 5: FOLLOW-UP ACTION LIST
    # ---------------------------------------------------------
    ws5 = wb.create_sheet("FOLLOW-UP ACTION LIST")
    headers5 = [
        "Customer Name", "Phone", "Reason", "Next Action", 
        "Due Date", "Assigned Agent", "Priority"
    ]
    ws5.append(headers5)
    
    for c in calls_data:
        analysis = c.get("analysis") or {}
        follow_ups = analysis.get("follow_up_recommendations") or []
        if follow_ups:
            for fu in follow_ups:
                if isinstance(fu, dict):
                    ws5.append(_safe_row([
                        c.get("customer_name", "Phone Lead"),
                        c.get("customer_id", ""),
                        fu.get("content", "Follow up"),
                        fu.get("type", "Call"),
                        "Tomorrow",
                        c.get("agent_name", "AI Agent"),
                        fu.get("priority", "Medium")
                    ]))
        else:
            action_items = analysis.get("action_items") or []
            for ai in action_items:
                ws5.append(_safe_row([
                    c.get("customer_name", "Phone Lead"),
                    c.get("customer_id", ""),
                    ai,
                    "Action Required",
                    "Tomorrow",
                    c.get("agent_name", "AI Agent"),
                    "Medium"
                ]))
    _apply_header_style(ws5)
    ws5.auto_filter.ref = ws5.dimensions
    ws5.freeze_panes = "A2"
    _auto_adjust_columns(ws5)

    # ---------------------------------------------------------
    # SHEET 6: LOST LEADS ANALYSIS
    # ---------------------------------------------------------
    ws6 = wb.create_sheet("LOST LEADS ANALYSIS")
    headers6 = [
        "Customer Name", "Reason Lost", "Budget Mismatch", "No Response", 
        "Competitor Chosen", "Not Interested", "Timeline Too Long"
    ]
    ws6.append(headers6)
    
    for c in calls_data:
        analysis = c.get("analysis") or {}
        if analysis.get("lead_temperature") == "Cold":
            objections = analysis.get("objections") or []
            obj_str = ", ".join([o.get("content", "") for o in objections if isinstance(o, dict)]) if objections else "Unknown"
            
            budget_mismatch = "Yes" if "budget" in obj_str.lower() or "price" in obj_str.lower() else "No"
            
            ws6.append(_safe_row([
                c.get("customer_name", "Phone Lead"),
                obj_str,
                budget_mismatch,
                "No",
                "No",
                "Yes" if "not interested" in obj_str.lower() else "No",
                "No"
            ]))
            
    _apply_header_style(ws6)
    ws6.auto_filter.ref = ws6.dimensions
    ws6.freeze_panes = "A2"
    _auto_adjust_columns(ws6)

    # Remove the initially created empty sheet
    if default_sheet.title == "Sheet":
        wb.remove(default_sheet)
        
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    
    return temp_file.name
