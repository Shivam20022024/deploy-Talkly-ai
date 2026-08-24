import base64
import json
from services.followup_service import followup_service
import mimetypes
import os
import re
import subprocess
import sys
import tempfile
import traceback
from datetime import datetime

try:
    from pydub import AudioSegment
    import shutil
    
    # HARDCODED FFMPEG CONFIGURATION (v8.1.1)
    FFMPEG_BIN_PATH = r"C:\Users\brijb\AppData\Local\Microsoft\WinGet\Packages\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe\ffmpeg-8.1.1-full_build\bin"
    AudioSegment.converter = os.path.join(FFMPEG_BIN_PATH, "ffmpeg.exe")
    AudioSegment.ffmpeg = os.path.join(FFMPEG_BIN_PATH, "ffmpeg.exe")
    AudioSegment.ffprobe = os.path.join(FFMPEG_BIN_PATH, "ffprobe.exe")
    
    # Startup Verification Logging
    print("--- FFMPEG STARTUP CONFIG ---")
    print("FFmpeg converter:", AudioSegment.converter)
    print("FFmpeg ffmpeg:", AudioSegment.ffmpeg)
    print("FFmpeg ffprobe:", AudioSegment.ffprobe)
    
    # Ensure bin directory is in PATH for subprocesses
    if os.path.exists(FFMPEG_BIN_PATH):
        os.environ["PATH"] = FFMPEG_BIN_PATH + os.pathsep + os.environ.get("PATH", "")
        print(f"Verified: FFmpeg directory prepended to PATH: {FFMPEG_BIN_PATH}")
    
    PYDUB_AVAILABLE = True
except ImportError:
    PYDUB_AVAILABLE = False
    print("WARNING: pydub not installed.")

import openpyxl
import requests
from dotenv import load_dotenv

# Load environment variables early
load_dotenv('.env.local', override=True)
load_dotenv(override=True)

# -----------------------------------------
# ENV HELPERS
# -----------------------------------------
def _env_bool(name, default=False):
    value = os.environ.get(name)
    if value is None: return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}

def _env_float(name, default):
    value = os.environ.get(name)
    if value is None: return default
    try: return float(value)
    except Exception: return default

def _now_ts():
    return datetime.utcnow().isoformat() + "Z"

# OPENROUTER CONFIG
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY")
OPENROUTER_STT_MODEL = "google/gemini-2.0-flash-001" # Verified ID
OPENROUTER_ANALYSIS_MODEL = os.environ.get("OPENROUTER_MODEL", "google/gemini-2.0-flash-001")
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"

# OPENAI CONFIG
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
OPENAI_STT_MODEL = "whisper-1"
OPENAI_ANALYSIS_MODEL = "gpt-4o"

STRICT_MODE = _env_bool("STRICT_MODE", True)

# -----------------------------------------
# CONFIG
# -----------------------------------------
TRANSCRIPT_DIR = "transcripts"
RESULTS_DIR = "results"
EXCEL_FILE = os.path.join(RESULTS_DIR, "analytics_results.xlsx")
CONVERTED_EXCEL_FILE = os.path.join(RESULTS_DIR, "converted_calls.xlsx")
SALES_CRM_FILE = os.path.join(RESULTS_DIR, "sales_crm.xlsx")

def _extract_json_object(text):
    if not text: return None
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start: return None
    candidate = text[start : end + 1]
    try: return json.loads(candidate)
    except Exception: return None

# -----------------------------------------
# PROMPT CONSTANT
# -----------------------------------------
ANALYSIS_SYSTEM_PROMPT = """
Analyze this real estate sales transcript and return a valid JSON object.
Focus heavily on identifying buyer interest and specific property requirements.
CRITICAL SCORING RULES:
1. If the customer states they are busy, hangs up quickly, or does not explicitly discuss property requirements, you MUST set the `lead_score` below 30, set `conversion_probability` below 10, and set `lead_temperature` to "Cold". Do not assume high intent without evidence.
2. Conversely, if the customer provides explicit requirements (e.g., budget, location, property type) AND states an immediate or very urgent timeline (e.g., "tomorrow", "immediately", "this week"), you MUST set `lead_score` above 80, set `conversion_probability` above 70, and set `lead_temperature` to "Hot".

The JSON schema must be:
{
  "summary": "Professional summary of the call",
  "sentiment": "positive|neutral|negative",
  "sentiment_confidence": 0.0-1.0,
  "sentiment_reason": "Brief explanation",
  "emotion": "dominant emotion",
  "detected_language": "Initial language detected (e.g., Hindi, English, Tamil, Hinglish)",
  "preferred_language": "Language preferred by customer later in the call (if any switch occurred)",
  "transcript_language": "Primary language of this transcript text",
  "intents": ["Purchase Inquiry", "Site Visit Request", "Price Negotiation"],
  "lead_temperature": "Hot|Warm|Cold",
  "lead_score": 85,
  "intent_label": "High Intent Buyer",
  "conversion_probability": 75,
  "customer_name": "John Doe",
  "property_requirements": {
    "budget": "80L - 1.2Cr",
    "location": "Sector 54",
    "propertyType": "Flat",
    "timeline": "3 months",
    "loanRequired": true,
    "investmentPurpose": "Self-Use"
  },
  "key_objections": [
    { "type": "Price", "severity": "Medium", "content": "Slightly over budget" }
  ],
  "follow_up_recommendations": [
    { "type": "WhatsApp", "priority": "High", "content": "Send brochure", "draft": "Hi John, here is the brochure..." }
  ],
  "action_items": ["Schedule site visit"],
  "agent_performance": {
    "talkRatio": 0.45,
    "interruptionCount": 2,
    "closingStrength": 8,
    "objectionHandlingScore": 7
  }
}

Transcript to analyze:
"""

# -----------------------------------------
# OPENROUTER HELPERS (Analysis Fallback Only)
# -----------------------------------------
def openrouter_analyze_transcript(transcript):
    if not transcript or not OPENROUTER_API_KEY: return None
    url = OPENROUTER_API_URL
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    
    prompt = ANALYSIS_SYSTEM_PROMPT + "\n\n" + transcript
    payload = {
        "model": OPENROUTER_ANALYSIS_MODEL,
        "messages": [
            {"role": "user", "content": prompt}
        ]
    }
    
    try:
        print(f"Analyzing via OpenRouter ({OPENROUTER_ANALYSIS_MODEL})...")
        response = requests.post(url, headers=headers, json=payload, timeout=90)
        if not response.ok:
            print(f"OpenRouter Analysis Error: {response.status_code} - {response.text}")
            return None
        
        data = response.json()
        content = data['choices'][0]['message']['content']
        return _extract_json_object(content)
    except Exception as e:
        print(f"OpenRouter Analysis CRITICAL FAILURE: {str(e)}")
        return None

# -----------------------------------------
# OPENAI HELPERS
# -----------------------------------------
def openai_transcribe_audio(audio_path):
    if not OPENAI_API_KEY: return ""
    
    file_size = os.path.getsize(audio_path)
    if file_size > 24 * 1024 * 1024: # 24MB limit to be safe
        print(f"File size {file_size/1024/1024:.2f}MB exceeds 25MB limit. Chunking...")
        return transcribe_large_audio(audio_path)

    url = "https://api.openai.com/v1/audio/transcriptions"
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}"}
    try:
        with open(audio_path, "rb") as f:
            files = {"file": (os.path.basename(audio_path), f)}
            data = {"model": OPENAI_STT_MODEL, "response_format": "text"}
            print(f"Transcribing via OpenAI ({OPENAI_STT_MODEL})...")
            response = requests.post(url, headers=headers, files=files, data=data, timeout=180)
            
            if not response.ok:
                print(f"OpenAI STT API Error: {response.status_code} - {response.text}")
                return ""
            
            return response.text.strip()
    except Exception as e:
        print(f"OpenAI Transcription CRITICAL FAILURE: {str(e)}")
        return ""

def transcribe_large_audio(audio_path):
    """
    Splits large audio into 10-minute chunks and transcribes each.
    """
    if not PYDUB_AVAILABLE:
        print("ERROR: pydub not installed. Cannot chunk large file. Falling back to OpenRouter...")
        return ""

    print(f"--- DEBUG CHUNKING START ---")
    print(f"Input file: {audio_path}")
    print(f"File exists: {os.path.exists(audio_path)}")
    print(f"FFmpeg converter: {getattr(AudioSegment, 'converter', 'Not Set')}")
    print(f"FFmpeg path: {getattr(AudioSegment, 'ffmpeg', 'Not Set')}")
    print(f"FFprobe path: {getattr(AudioSegment, 'ffprobe', 'Not Set')}")

    try:
        # Explicitly check if ffprobe/ffmpeg are actually reachable
        try:
            subprocess.run([AudioSegment.converter, "-version"], capture_output=True, check=True)
            print("FFmpeg verification: SUCCESS")
        except Exception as e:
            print(f"FFmpeg verification: FAILED - {str(e)}")

        print(f"Loading audio segment from {audio_path}...")
        audio = AudioSegment.from_file(audio_path)
        print(f"Audio loaded. Duration: {len(audio)/1000/60:.2f} minutes")
        
        ten_minutes = 10 * 60 * 1000 # pydub works in milliseconds
        chunks = [audio[i:i + ten_minutes] for i in range(0, len(audio), ten_minutes)]
        
        full_transcript = []
        base_name = os.path.splitext(os.path.basename(audio_path))[0]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            print(f"Created temporary directory: {tmpdir}")
            for i, chunk in enumerate(chunks):
                chunk_path = os.path.join(tmpdir, f"{base_name}_part{i}.mp3")
                print(f"Processing chunk {i+1}/{len(chunks)}: {chunk_path}")
                chunk.export(chunk_path, format="mp3", bitrate="128k")
                
                if not os.path.exists(chunk_path):
                    print(f"ERROR: Chunk export failed for part {i+1}")
                    continue

                # Use the standard transcription function for the chunk
                url = "https://api.openai.com/v1/audio/transcriptions"
                headers = {"Authorization": f"Bearer {OPENAI_API_KEY}"}
                with open(chunk_path, "rb") as f:
                    files = {"file": (os.path.basename(chunk_path), f)}
                    data = {"model": OPENAI_STT_MODEL, "response_format": "text"}
                    print(f"Transcribing chunk {i+1} via OpenAI...")
                    resp = requests.post(url, headers=headers, files=files, data=data, timeout=120)
                    if resp.ok:
                        full_transcript.append(resp.text.strip())
                    else:
                        print(f"Chunk {i+1} failed: {resp.status_code} - {resp.text}")
            
        print(f"--- DEBUG CHUNKING COMPLETE ---")
        return " ".join(full_transcript)
    except Exception as e:
        print(f"--- DEBUG CHUNKING FAILURE ---")
        print(f"Error Type: {type(e).__name__}")
        print(f"Error Message: {str(e)}")
        print(f"Traceback:")
        traceback.print_exc()
        if "ffprobe" in str(e).lower() or "ffmpeg" in str(e).lower() or "[WinError 2]" in str(e):
            print("CRITICAL: FFmpeg/FFprobe binaries are missing or unreachable despite configuration.")
            print(f"Current PATH: {os.environ.get('PATH', '')}")
        return ""

def openai_analyze_transcript(transcript):
    if not transcript or not OPENAI_API_KEY: return None
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": OPENAI_ANALYSIS_MODEL,
        "messages": [{"role": "user", "content": ANALYSIS_SYSTEM_PROMPT + transcript}],
        "response_format": {"type": "json_object"}
    }
    try:
        print(f"Analyzing via OpenAI ({OPENAI_ANALYSIS_MODEL})...")
        response = requests.post(url, headers=headers, json=payload, timeout=90)
        if not response.ok:
            print(f"OpenAI Analysis API Error: {response.status_code} - {response.text}")
            return None
        return _extract_json_object(response.json()["choices"][0]["message"]["content"])
    except Exception as e:
        print(f"OpenAI Analysis CRITICAL FAILURE: {str(e)}")
        return None

def openai_refine_transcript(transcript):
    if not transcript or not OPENAI_API_KEY: return transcript
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    prompt = f"Identify distinct speakers in this raw transcript based on context and assign appropriate labels (e.g., 'Agent', 'Customer', 'Third Party', 'Manager', etc.). Return ONLY the labeled conversation.\n\nRaw:\n{transcript}"
    payload = {"model": OPENAI_ANALYSIS_MODEL, "messages": [{"role": "user", "content": prompt}]}
    try:
        print(f"Refining transcript for speakers...")
        response = requests.post(url, headers=headers, json=payload, timeout=90)
        if not response.ok: return transcript
        return response.json()["choices"][0]["message"]["content"].strip()
    except Exception: return transcript

# -----------------------------------------
# EXCEL & FILE HELPERS
# -----------------------------------------
def get_weekly_excel_file():
    now = datetime.utcnow()
    y, w, _ = now.isocalendar()
    return os.path.join("results", f"weekly_calls_{y}_W{w}.xlsx")

def safe_write(path, text):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f: f.write(text or "")
    return path

def write_excel(path, row):
    os.makedirs("results", exist_ok=True)
    new_file = not os.path.exists(path)
    wb = openpyxl.Workbook() if new_file else openpyxl.load_workbook(path)
    ws = wb.active
    if new_file: ws.append(list(row.keys()))
    ws.append(list(row.values()))
    wb.save(path)

def export_calls_to_excel(calls, path):
    os.makedirs("results", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead Intelligence Report"
    
    headers = [
        "Date", "Customer Name", "Detected Language", "Preferred Language", "Transcript Language", "Lead Temperature", 
        "Lead Score (%)", "Conversion Prob (%)", "Budget", "Location", 
        "Property Type", "Timeline", "Summary", "Sentiment", "Action Items", "Call ID"
    ]
    ws.append(headers)
    
    for call in calls:
        dt = call.get("created_at")
        dt_str = dt.strftime("%Y-%m-%d %H:%M:%S") if isinstance(dt, datetime) else str(dt)
        reqs = call.get("analysis", {}).get("property_requirements") or {}
        
        row = [
            dt_str,
            call.get("customer_name") or "Phone Lead",
            call.get("detected_language") or call.get("language") or "English",
            call.get("preferred_language") or call.get("detected_language") or call.get("language") or "English",
            call.get("transcript_language") or call.get("detected_language") or call.get("language") or "English",
            call.get("analysis", {}).get("lead_temperature") or "Warm",
            call.get("analysis", {}).get("lead_score") or call.get("analysis", {}).get("intent_score") or 0,
            call.get("analysis", {}).get("conversion_probability") or 0,
            reqs.get("budget") or "N/A",
            reqs.get("location") or "N/A",
            reqs.get("propertyType") or "N/A",
            reqs.get("timeline") or "N/A",
            call.get("summary") or "",
            call.get("sentiment") or "neutral",
            ", ".join(call.get("analysis", {}).get("action_items", [])),
            call.get("call_id")
        ]
        ws.append(row)
        
    wb.save(path)
    return path

# -----------------------------------------
# MAIN PROCESSING
# -----------------------------------------
def analyze_transcript_text(transcript):
    if not transcript:
        return {
            "summary": "No transcript available for analysis.",
            "sentiment": "neutral",
            "intents": [],
            "detected_language": "Unknown",
            "preferred_language": "Unknown",
            "transcript_language": "Unknown",
            "property_requirements": {},
            "action_items": []
        }

    result = None
    if OPENAI_API_KEY:
        result = openai_analyze_transcript(transcript)
    
    if not result and OPENROUTER_API_KEY:
        result = openrouter_analyze_transcript(transcript)
    
    if not result:
        return {
            "summary": "AI Analysis failed to process the transcript.",
            "sentiment": "neutral",
            "intents": [],
            "detected_language": "Unknown",
            "preferred_language": "Unknown",
            "transcript_language": "Unknown",
            "property_requirements": {},
            "action_items": []
        }
    
    # Enforce Lead Temperature based on Score for consistency
    score = result.get("lead_score") or result.get("intent_score") or 0
    if isinstance(score, (int, float)):
        # Ensure both fields are populated so the frontend always finds the score
        result["lead_score"] = score
        result["intent_score"] = score
        
        if score >= 80:
            result["lead_temperature"] = "Hot"
        elif score >= 40:
            result["lead_temperature"] = "Warm"
        else:
            result["lead_temperature"] = "Cold"
            
    return result

def process_uploaded_audio(audio_path):
    print(f"--- STARTING PROCESSING FOR: {audio_path} ---")
    filename = os.path.basename(audio_path)
    base = os.path.splitext(filename)[0]

    raw = ""
    if OPENAI_API_KEY: 
        raw = openai_transcribe_audio(audio_path)
    
    if not raw:
        print(f"ERROR: Transcription failed for {audio_path}. Check OpenAI API or FFmpeg configuration.")
        return {
            "summary": "Transcription failed.",
            "sentiment": "neutral",
            "intents": [],
            "language_detected": "Unknown",
            "property_requirements": {},
            "action_items": []
        }
    
    transcript = raw
    if OPENAI_API_KEY and raw:
        refined = openai_refine_transcript(raw)
        if refined: transcript = refined

    result = analyze_transcript_text(transcript)
    
    summary = result.get("summary") or "Summary not available"
    sentiment = result.get("sentiment") or "neutral"
    intents = result.get("intents") or []
    detected_language = result.get("detected_language") or result.get("language_detected") or "English"
    preferred_language = result.get("preferred_language") or detected_language
    transcript_language = result.get("transcript_language") or detected_language
    
    customer_name = result.get("customer_name") or "Phone Lead"
    reqs = result.get("property_requirements") or {}
    budget = reqs.get("budget") or "N/A"
    location = reqs.get("location") or "N/A"
    prop_type = reqs.get("propertyType") or "N/A"
    timeline = reqs.get("timeline") or "N/A"
    
    lead_score = result.get("lead_score") or result.get("intent_score") or 0
    temp = result.get("lead_temperature") or "Warm"
    conv_prob = result.get("conversion_probability") or 0
    action_items = ", ".join(result.get("action_items", []))
    key_objections = result.get("key_objections", [])

    safe_write(os.path.join("transcripts", base + ".txt"), transcript)

    row = {
        "Date": _now_ts(),
        "Customer Name": customer_name,
        "Language": detected_language,
        "Preferred Language": preferred_language,
        "Transcript Language": transcript_language,
        "Lead Temperature": temp,
        "Intent Score (%)": lead_score,
        "Conversion Prob (%)": conv_prob,
        "Budget": budget,
        "Location": location,
        "Property Type": prop_type,
        "Timeline": timeline,
        "Summary": summary,
        "Sentiment": sentiment,
        "Action Items": action_items,
        "Lead Score": lead_score,
        "File Name": filename
    }

    try:
        write_excel("results/analytics_results.xlsx", row)
        write_excel(get_weekly_excel_file(), row)
    except Exception as e: print(f"Excel logging failed: {e}")

    result_dict = {
        "call_id": base,
        "transcript": transcript,
        "summary": summary,
        "sentiment": sentiment,
        "intents": intents,
        "language": detected_language,
        "detected_language": detected_language,
        "preferred_language": preferred_language,
        "transcript_language": transcript_language,
        "analysis": result or {}
    }

    print("Drafting follow-up email...")
    draft = followup_service.generate_followup_draft(result_dict)
    if draft:
        result_dict["analysis"]["followup_draft"] = draft
        print("Follow-up email drafted successfully.")

    return result_dict

if __name__ == "__main__":
    print("process_audio.py ready")
