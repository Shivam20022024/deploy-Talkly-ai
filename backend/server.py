# server.py
import os
import sys
import warnings

# Suppress pydub RuntimeWarnings
warnings.filterwarnings("ignore", category=RuntimeWarning, module="pydub")

# Configure FFmpeg paths at the very start to suppress pydub warnings
USER_FFMPEG_BIN = r"C:\Users\brijb\AppData\Local\Microsoft\WinGet\Packages\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe\ffmpeg-8.1-full_build\bin"
if os.path.exists(USER_FFMPEG_BIN):
    os.environ["PATH"] += os.pathsep + USER_FFMPEG_BIN

import time
import json
import uuid
import asyncio
import requests
import mimetypes
import subprocess
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException, Body, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
import io
import openpyxl
from fastapi.concurrency import run_in_threadpool
from process_audio import process_uploaded_audio, get_weekly_excel_file, analyze_transcript_text
import mongodb
from services.language_service import LanguageService
from services.intelligence_service import IntelligenceService
language_service = LanguageService()
intelligence_service = IntelligenceService()
from fastapi.responses import FileResponse
from dotenv import load_dotenv

from contextlib import asynccontextmanager

# Load env values
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env.local"))

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup logic
    await mongodb.ensure_indexes()
    print("--- PROVIDER CONFIG ---")
    print(f"OpenRouter Analysis: {os.environ.get('OPENROUTER_MODEL', 'Not Set')} (Key: {bool(os.environ.get('OPENROUTER_API_KEY'))})")
    print(f"OpenAI Analysis: gpt-4o (Key: {bool(os.environ.get('OPENAI_API_KEY'))})")
    yield
    # Shutdown logic (if any)

app = FastAPI(lifespan=lifespan)

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from inbound.routes import inbound_router
app.include_router(inbound_router, prefix="/api/v1/inbound")

@app.get("/")
@app.head("/")
async def health_check():
    return {"status": "ok", "service": "TalklyAI Backend"}

@app.get("/api/v1/analytics/dashboard")
async def get_analytics_dashboard():
    db = mongodb.get_db()
    metrics = await intelligence_service.get_dashboard_metrics(db)
    return {"status": "success", "data": metrics}

@app.get("/api/v1/customers/{customer_id}/timeline")
async def get_customer_timeline(customer_id: str):
    db = mongodb.get_db()
    timeline = await intelligence_service.get_customer_timeline(db, customer_id)
    return {"status": "success", "data": timeline}

BASE_URL = os.environ.get("BASE_URL", "http://localhost:8000")

# Startup logic moved to lifespan context manager above

@app.get("/calls")
async def get_calls(limit: int = 20, skip: int = 0):
    db = mongodb.get_db()
    cursor = db.calls.find().sort("created_at", -1).skip(skip).limit(limit)
    calls = []
    async for d in cursor:
        d["_id"] = None
        calls.append(d)
    return calls

@app.get("/calls/{call_id}")
async def get_call(call_id: str):
    db = mongodb.get_db()
    doc = await db.calls.find_one({"call_id": call_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Call not found")
    doc["_id"] = None
    return doc

@app.post("/calls/trigger")
async def trigger_bolna_call(payload: dict = Body(...)):
    phone_number = payload.get("phone_number")
    lead_id = payload.get("lead_id")
    campaign_language = payload.get("campaign_language", "English")
    ai_voice = payload.get("ai_voice", "Default")
    voice_gender = payload.get("voice_gender", "Female")
    regional_accent = payload.get("regional_accent", "Default")

    api_key = os.environ.get("BOLNA_API_KEY", "").strip()
    
    db = mongodb.get_db()
    
    # 2. Dynamically fetch Agent ID based on Database mapping
    agent_id = await language_service.get_agent_id_for_language(db, campaign_language)

    if not api_key or not agent_id:
        raise HTTPException(status_code=400, detail="Bolna.ai API Key or Agent ID not configured")

    
    # Always fetch latest BASE_URL to ensure it reflects current ngrok/tunnel
    current_base_url = os.environ.get("BASE_URL", "http://localhost:8000").strip()
    
    await db.calls.insert_one({
        "call_id": lead_id,
        "customer_id": phone_number,
        "customer_name": "Phone Lead",
        "direction": "outbound",
        "transcript": "",
        "status": "Initiating",
        "created_at": datetime.utcnow(),
        "language": campaign_language,
        "detected_language": campaign_language,
        "preferred_language": campaign_language,
        "transcript_language": campaign_language,
        "ai_voice": ai_voice,
        "voice_gender": voice_gender,
        "regional_accent": regional_accent
    })

    url = "https://api.bolna.ai/call"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    bolna_payload = {
        "agent_id": agent_id,
        "recipient_phone_number": phone_number,
        "webhook_url": f"{current_base_url}/webhooks/bolna",
        "user_data": {
            "lead_id": lead_id,
            "customer_name": "Phone Lead",
            "agent_name": "AI Agent",
            "campaign_language": campaign_language,
            "ai_voice": ai_voice,
            "voice_gender": voice_gender,
            "regional_accent": regional_accent
        }
    }

    if campaign_language.lower() != "english":
        bolna_payload["agent_config"] = {
            "prompt": {
                "system_prompt": f"You are an AI sales agent for TalklyAI. You MUST speak entirely in the {campaign_language} language. Please reply naturally in {campaign_language}."
            }
        }

    try:
        print(f"Triggering Bolna call to {phone_number} with webhook_url: {bolna_payload['webhook_url']}")
        response = requests.post(url, json=bolna_payload, headers=headers, timeout=30)
        
        if not response.ok:
            await db.calls.update_one({"call_id": lead_id}, {"$set": {"status": "Failed"}})
            raise HTTPException(status_code=response.status_code, detail=f"Bolna Error: {response.text}")
        
        bolna_data = response.json()
        execution_id = bolna_data.get("execution_id")
        
        if execution_id:
            # Start a background polling task as a fallback for webhooks
            asyncio.create_task(poll_bolna_execution(execution_id, lead_id, api_key))
            
        return bolna_data
    except HTTPException:
        raise
    except Exception as e:
        await db.calls.update_one({"call_id": lead_id}, {"$set": {"status": "Error"}})
        raise HTTPException(status_code=500, detail=f"API Error: {str(e)}")

async def process_bulk_calls(phone_numbers: list, campaign_language: str = "English", ai_voice: str = "Default", voice_gender: str = "Female", regional_accent: str = "Default"):
    api_key = os.environ.get("BOLNA_API_KEY", "").strip()
    
    db = mongodb.get_db()
    
    # Dynamically fetch Agent ID based on Database mapping
    agent_id = await language_service.get_agent_id_for_language(db, campaign_language)
    
    if not api_key or not agent_id:
        return
    current_base_url = os.environ.get("BASE_URL", "http://localhost:8000").strip()
    url = "https://api.bolna.ai/call"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    for idx, phone in enumerate(phone_numbers):
        phone_str = str(phone).strip()
        if not phone_str:
            continue
            
        lead_id = f"bulk_{int(time.time() * 1000)}_{idx}"
        
        await db.calls.insert_one({
            "call_id": lead_id,
            "customer_id": phone_str,
            "customer_name": "Bulk Phone Lead",
            "direction": "outbound",
            "transcript": "",
            "status": "Initiating",
            "created_at": datetime.utcnow(),
            "language": campaign_language,
            "detected_language": campaign_language,
            "preferred_language": campaign_language,
            "transcript_language": campaign_language,
            "ai_voice": ai_voice,
            "voice_gender": voice_gender,
            "regional_accent": regional_accent
        })
        
        bolna_payload = {
            "agent_id": agent_id,
            "recipient_phone_number": phone_str,
            "webhook_url": f"{current_base_url}/webhooks/bolna",
            "user_data": {
                "lead_id": lead_id,
                "customer_name": "Bulk Phone Lead",
                "agent_name": "AI Agent",
                "campaign_language": campaign_language,
                "ai_voice": ai_voice,
                "voice_gender": voice_gender,
                "regional_accent": regional_accent
            }
        }
        
        if campaign_language.lower() != "english":
            bolna_payload["agent_config"] = {
                "prompt": {
                    "system_prompt": f"You are an AI sales agent for TalklyAI. You MUST speak entirely in the {campaign_language} language. Please reply naturally in {campaign_language}."
                }
            }
        
        try:
            print(f"Triggering Bulk Bolna call to {phone_str} with webhook_url: {bolna_payload['webhook_url']}")
            # Run blocking request in threadpool
            response = await run_in_threadpool(requests.post, url, json=bolna_payload, headers=headers, timeout=30)
            
            if response.ok:
                bolna_data = response.json()
                execution_id = bolna_data.get("execution_id")
                if execution_id:
                    asyncio.create_task(poll_bolna_execution(execution_id, lead_id, api_key))
            else:
                await db.calls.update_one({"call_id": lead_id}, {"$set": {"status": "Failed"}})
                print(f"Bolna Bulk Error for {phone_str}: {response.text}")
        except HTTPException:
            raise
        except Exception as e:
            print(f"Error processing bulk call for {phone_str}: {e}")
            await db.calls.update_one({"call_id": lead_id}, {"$set": {"status": "Error"}})
            print(f"API Bulk Error for {phone_str}: {str(e)}")
            
        await asyncio.sleep(0.5)

@app.post("/calls/trigger-bulk")
async def trigger_bolna_bulk(
    background_tasks: BackgroundTasks, 
    file: UploadFile = File(...),
    campaign_language: str = Form("English"),
    ai_voice: str = Form("Default"),
    voice_gender: str = Form("Female"),
    regional_accent: str = Form("Default")
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = wb.active
        
        phone_numbers = []
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        phone_col_idx = 0
        
        if header_row:
            for idx, cell in enumerate(header_row):
                if cell and str(cell).lower() in ["phone", "phone number", "phone_number", "number", "mobile"]:
                    phone_col_idx = idx
                    break
                    
        for row in sheet.iter_rows(min_row=2 if header_row else 1, values_only=True):
            if row and len(row) > phone_col_idx and row[phone_col_idx]:
                phone_numbers.append(row[phone_col_idx])
                
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
        
    if not phone_numbers:
        raise HTTPException(status_code=400, detail="No phone numbers found in the Excel file.")
        
    background_tasks.add_task(process_bulk_calls, phone_numbers, campaign_language, ai_voice, voice_gender, regional_accent)
    
    return {
        "status": "success",
        "message": f"Successfully initiated bulk calls for {len(phone_numbers)} numbers.",
        "count": len(phone_numbers)
    }

@app.post("/calls/send-email")
async def send_email(payload: dict = Body(...)):
    to_email = payload.get("to")
    subject = payload.get("subject")
    body = payload.get("body")
    
    if not to_email:
        raise HTTPException(status_code=400, detail="Missing 'to' email address.")
        
    print(f"--- Simulating Email Send ---")
    print(f"To: {to_email}")
    print(f"Subject: {subject}")
    print(f"Body Snippet: {str(body)[:100]}...")
    print(f"-----------------------------")
    
    # Simulate network delay
    await asyncio.sleep(1.5)
    
    return {"status": "success", "message": "Email simulated successfully."}


async def poll_bolna_execution(execution_id: str, lead_id: str, api_key: str):
    """
    Polls Bolna API for transcript updates in case webhooks fail.
    """
    db = mongodb.get_db()
    url = f"https://api.bolna.ai/executions/{execution_id}"
    headers = {"Authorization": f"Bearer {api_key}"}
    
    log_msg = f"Starting polling fallback for Execution {execution_id} (Lead: {lead_id})"
    print(log_msg)
    with open("bolna_polling.log", "a") as f:
        f.write(f"{datetime.utcnow().isoformat()} - {log_msg}\n")
    
    # Poll for up to 10 minutes
    for i in range(60): 
        try:
            await asyncio.sleep(10)
            # Use run_in_threadpool for blocking requests.get
            response = await run_in_threadpool(requests.get, url, headers=headers, timeout=15)
            
            if response.ok:
                # Force UTF-8 decoding for Hindi/International characters
                response.encoding = 'utf-8'
                data = response.json()
                transcript = data.get("transcript") or ""
                status = data.get("status", "Active")
                
                with open("bolna_polling.log", "a") as f:
                    f.write(f"{datetime.utcnow().isoformat()} - Poll {i}: Status={status}, TranscriptLen={len(transcript)}\n")

                if transcript:
                    update_data = {
                        "transcript": transcript,
                        "last_poll_at": datetime.utcnow().isoformat() + "Z"
                    }
                    
                    if status.lower() in ["completed", "finished", "ended"]:
                        print(f"Polling: Call {lead_id} completed. Triggering analysis...")
                        analysis_result = await run_in_threadpool(analyze_transcript_text, transcript)
                        if analysis_result:
                            update_data.update({
                                "status": "Analyzed",
                                "summary": analysis_result.get("summary"),
                                "sentiment": analysis_result.get("sentiment"),
                                "analysis": analysis_result
                            })
                        else:
                            update_data["status"] = "Completed"
                        
                        await db.calls.update_one({"call_id": lead_id}, {"$set": update_data})
                        break # Stop polling
                    else:
                        update_data["status"] = "Active"
                        await db.calls.update_one({"call_id": lead_id}, {"$set": update_data})
            else:
                with open("bolna_polling.log", "a") as f:
                    f.write(f"{datetime.utcnow().isoformat()} - Poll {i}: API Error {response.status_code}\n")
                    
        except Exception as e:
            err_msg = f"Polling error for {lead_id}: {str(e)}"
            print(err_msg)
            with open("bolna_polling.log", "a") as f:
                f.write(f"{datetime.utcnow().isoformat()} - {err_msg}\n")
    
    print(f"Finished polling fallback for {lead_id}")

@app.post("/webhooks/bolna")
async def bolna_webhook(data: dict = Body(...)):
    db = mongodb.get_db()
    
    # Save to log file
    with open("bolna_webhook.log", "a") as f:
        f.write(f"{datetime.utcnow().isoformat()} - {json.dumps(data)}\n")
        
    print(f"WEBHOOK RECEIVED FROM BOLNA: {json.dumps(data)[:300]}...")
    
    user_data = data.get("user_data") or {}
    if isinstance(user_data, str):
        try: user_data = json.loads(user_data)
        except: user_data = {}
            
    lead_id = user_data.get("lead_id") or data.get("lead_id") or data.get("call_id")
    transcript = data.get("transcript") or data.get("conversation_history") or data.get("telephony_data", {}).get("transcript", "")
    status = data.get("status", "Active")
    
    if not lead_id:
        phone = data.get("recipient_phone_number") or data.get("telephony_data", {}).get("to_number")
        if phone:
            phone_clean = phone.replace('+', '').strip()
            potential_call = await db.calls.find_one({
                "customer_id": {"$regex": f"{phone_clean}$"}, 
                "status": {"$in": ["Initiating", "Active"]}
            })
            if potential_call:
                lead_id = potential_call["call_id"]
                
        # If STILL no lead_id, it's a completely new inbound call!
        if not lead_id:
            customer_phone = data.get("telephony_data", {}).get("from_number") or data.get("caller_phone_number") or "Unknown Inbound"
            lead_id = data.get("call_id") or data.get("execution_id") or f"inbound_{int(time.time()*1000)}"
            
            await db.calls.insert_one({
                "call_id": lead_id,
                "customer_id": customer_phone,
                "customer_name": "Inbound Caller",
                "direction": "inbound",
                "transcript": "",
                "status": "Active",
                "created_at": datetime.utcnow(),
                "language": "Detecting...",
                "detected_language": "Unknown",
                "preferred_language": "Unknown",
                "transcript_language": "Unknown"
            })
            print(f"Created new INBOUND lead record: {lead_id} from {customer_phone}")

    if lead_id:
        update_data = {
            "transcript": transcript,
            "last_webhook_at": datetime.utcnow().isoformat() + "Z"
        }
        
        if status.lower() in ["completed", "finished", "ended"]:
            print(f"Call {lead_id} completed. Triggering AI Analysis...")
            try:
                analysis_result = await run_in_threadpool(analyze_transcript_text, transcript)
                if analysis_result:
                    detected = analysis_result.get("detected_language") or analysis_result.get("language_detected") or "English"
                    update_data.update({
                        "status": "Analyzed",
                        "summary": analysis_result.get("summary"),
                        "sentiment": analysis_result.get("sentiment"),
                        "language": detected,
                        "detected_language": detected,
                        "preferred_language": analysis_result.get("preferred_language") or detected,
                        "transcript_language": analysis_result.get("transcript_language") or detected,
                        "analysis": analysis_result
                    })
                else:
                    update_data["status"] = "Completed"
            except Exception as e:
                print(f"Analysis failed for {lead_id}: {str(e)}")
                update_data["status"] = "Completed"
        else:
            update_data["status"] = "Active"

        await db.calls.update_one({"call_id": lead_id}, {"$set": update_data})
    return {"status": "success"}

@app.post("/process-audio")
async def process_audio_api(
    file: UploadFile = File(...), 
    agent_name: str = Form("AI Agent"),
    employee_id: str = Form(None),
    employee_email: str = Form(None)
):
    temp_path = None
    try:
        timestamp = int(time.time() * 1000)
        temp_path = f"temp_{timestamp}_{file.filename}"
        with open(temp_path, "wb") as f:
            f.write(await file.read())
        
        result = await run_in_threadpool(process_uploaded_audio, temp_path)
        now = datetime.utcnow()
        unique_call_id = f"call_{timestamp}"
        doc = {
            "call_id": unique_call_id,
            "customer_id": result.get("analysis", {}).get("customer_name") or "Phone Lead",
            "customer_name": result.get("analysis", {}).get("customer_name") or "Phone Lead",
            "direction": "outbound",
            "agent_name": agent_name,
            "employee_id": employee_id,
            "employee_email": employee_email,
            "sentiment": str(result.get("sentiment", "neutral")).lower(),
            "summary": result.get("summary"),
            "transcript": result.get("transcript"),
            "analysis": result.get("analysis", {}),
            "created_at": now,
            "status": "Analyzed"
        }
        db = mongodb.get_db()
        await db.calls.insert_one(doc)
        doc["_id"] = None
        if temp_path and os.path.exists(temp_path): os.remove(temp_path)
        return doc
    except Exception as e:
        print(f"ERROR PROCESSING AUDIO: {str(e)}")
        if temp_path and os.path.exists(temp_path): os.remove(temp_path)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{report_type}")
async def download_report(report_type: str):
    if report_type == "excel":
        file_path = os.path.join(BASE_DIR, "results", "analytics_results.xlsx")
        if os.path.exists(file_path):
            return FileResponse(file_path, filename="analytics_results.xlsx")
        return {"status": "error", "message": "Report not found"}
    # Mock download for demo
    return {"status": "success", "message": f"{report_type} report ready"}

# --- Phase 3 CRM & Follow-Up Endpoints ---

from services.crm_service import crm_service

@app.post("/api/v1/calls/{call_id}/sync-crm")
async def sync_call_to_crm(call_id: str):
    try:
        db = mongodb.get_db()
        await crm_service.sync_call_to_crm(db, call_id)
        return {"status": "success", "message": "Successfully synced to CRM."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/v1/calls/{call_id}/send-followup")
async def send_followup_email(call_id: str):
    # Simulates sending an email
    try:
        db = mongodb.get_db()
        import asyncio
        await asyncio.sleep(1) # Simulate network delay
        
        # Mark email as sent in DB
        await db.calls.update_one(
            {"call_id": call_id},
            {"$set": {
                "followup_sent": True,
                "followup_sent_at": datetime.utcnow().isoformat()
            }}
        )
        return {"status": "success", "message": "Email sent successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
